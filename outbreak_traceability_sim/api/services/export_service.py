"""
Export service for generating Excel files from simulation data.

Exports all transactional (CTE/KDE) and case data for audit purposes.
"""

from datetime import datetime
from io import BytesIO
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .simulation_service import simulation_service, SimulationRun


def _style_header_row(ws, num_cols: int):
    """Apply header styling to first row."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border


def _auto_column_width(ws):
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column].width = adjusted_width


def _get_node_name(network, node_id: UUID) -> str:
    """Get human-readable name for a node."""
    node = network.get_node(node_id)
    if not node:
        return str(node_id)[:8]

    if hasattr(node, 'farm_name'):
        return node.farm_name
    elif hasattr(node, 'facility_name'):
        return node.facility_name
    elif hasattr(node, 'store_name'):
        return f"{node.store_name} #{node.store_number}"
    return str(node_id)[:8]


def _get_node_type(network, node_id: UUID) -> str:
    """Get node type string."""
    node = network.get_node(node_id)
    if not node:
        return "unknown"
    return node.node_type.value


def export_simulation_to_excel(simulation_id: str) -> Optional[BytesIO]:
    """
    Export simulation data to Excel workbook.

    Args:
        simulation_id: ID of the simulation to export

    Returns:
        BytesIO buffer containing Excel file, or None if simulation not found
    """
    run = simulation_service.get_run(simulation_id)
    if not run or not run.simulator:
        return None

    simulator = run.simulator
    network = simulator.network
    flow_sim = simulator.flow_sim
    seeder = simulator.seeder
    cases = simulator.cases or []
    lot_graph = simulator.lot_graph

    if not network or not flow_sim:
        return None

    # Build contamination lookup from seeder's propagation data
    contamination_lookup: dict[str, float] = {}
    if seeder:
        contamination_lookup = seeder.contamination_propagation.copy()
        # Also include source TLCs
        for tlc in seeder.contaminated_source_tlcs:
            if tlc not in contamination_lookup:
                contamination_lookup[tlc] = 1.0

    def is_tlc_contaminated(tlc: str) -> tuple[bool, float]:
        """Check if a TLC is contaminated and return probability."""
        if tlc in contamination_lookup:
            return True, contamination_lookup[tlc]
        # Also check lot_graph
        if lot_graph and tlc in lot_graph.lots:
            lot = lot_graph.lots[tlc]
            if lot.is_contaminated or lot.contamination_probability > 0:
                return True, lot.contamination_probability
        return False, 0.0

    def check_shipment_contamination(tlcs: list[str]) -> tuple[bool, float]:
        """Check if any TLCs in a shipment are contaminated."""
        max_prob = 0.0
        is_contaminated = False
        for tlc in tlcs:
            contaminated, prob = is_tlc_contaminated(tlc)
            if contaminated:
                is_contaminated = True
                max_prob = max(max_prob, prob)
        return is_contaminated, max_prob

    wb = Workbook()

    # =========================================================================
    # Sheet 1: Metadata
    # =========================================================================
    ws_meta = wb.active
    ws_meta.title = "Metadata"

    ws_meta.append(["Parameter", "Value"])
    ws_meta.append(["Simulation ID", simulation_id])
    ws_meta.append(["Export Generated At", datetime.now().isoformat()])
    ws_meta.append(["Random Seed", run.config.get("random_seed", "None")])
    ws_meta.append([])
    ws_meta.append(["--- Network Configuration ---", ""])
    ws_meta.append(["Number of Farms", run.config.get("num_farms", "")])
    ws_meta.append(["Number of Packers", run.config.get("num_packers", "")])
    ws_meta.append(["Number of Distribution Centers", run.config.get("num_distribution_centers", "")])
    ws_meta.append(["Number of Retailers", run.config.get("num_retailers", "")])
    ws_meta.append(["Retailers with Delis %", run.config.get("retailers_with_delis_pct", "")])
    ws_meta.append([])
    ws_meta.append(["--- Simulation Period ---", ""])
    ws_meta.append(["Simulation Days", run.config.get("simulation_days", "")])
    ws_meta.append(["Start Date", simulator.config.start_date.isoformat()])
    ws_meta.append(["End Date", simulator.config.end_date.isoformat()])
    ws_meta.append([])
    ws_meta.append(["--- Contamination ---", ""])
    ws_meta.append(["Pathogen", run.config.get("pathogen", "")])
    ws_meta.append(["Contamination Rate", run.config.get("contamination_rate", "")])
    ws_meta.append(["Contamination Duration Days", run.config.get("contamination_duration_days", "")])
    ws_meta.append(["Contamination Source Farm", _get_node_name(network, simulator.contaminated_farm_id) if simulator.contaminated_farm_id else ""])
    ws_meta.append(["Contamination Source Farm ID", str(simulator.contaminated_farm_id) if simulator.contaminated_farm_id else ""])
    ws_meta.append([])
    ws_meta.append(["--- Investigation Parameters ---", ""])
    ws_meta.append(["Interview Success Rate", run.config.get("interview_success_rate", "")])
    ws_meta.append(["Record Collection Window Days", run.config.get("record_collection_window_days", "")])
    ws_meta.append([])
    ws_meta.append(["--- Summary Statistics ---", ""])
    ws_meta.append(["Total Lots Created", flow_sim.total_lots_created])
    ws_meta.append(["Total Shipments", len(flow_sim.shipments)])
    ws_meta.append(["Total Cases", len(cases)])
    ws_meta.append(["Deterministic Lot Links", flow_sim.deterministic_lot_links])
    ws_meta.append(["Probabilistic Lot Links", flow_sim.probabilistic_lot_links])

    if seeder:
        ws_meta.append(["Contaminated Source TLCs", len(seeder.contaminated_source_tlcs)])
        ws_meta.append(["Contaminated Downstream TLCs", len(seeder.contamination_propagation)])

    _style_header_row(ws_meta, 2)
    _auto_column_width(ws_meta)
    ws_meta.freeze_panes = 'A2'

    # =========================================================================
    # Sheet 2: Network Nodes
    # =========================================================================
    ws_nodes = wb.create_sheet("Network_Nodes")

    headers = [
        "Node ID", "Node Type", "Name", "City", "State",
        "Is Contamination Source", "Received Contaminated Product"
    ]
    ws_nodes.append(headers)

    # Get contamination info
    contaminated_node_ids = set()
    if seeder:
        for event in seeder.contamination_events:
            contaminated_node_ids.add(str(event.farm_id))

    # Track nodes that received contaminated product
    # Check both the shipment flag AND the TLCs in the shipment against contamination lookup
    nodes_with_contamination = set()
    for shipment in flow_sim.shipments:
        # Check if shipment is flagged as contaminated
        if shipment.contains_contaminated_product:
            nodes_with_contamination.add(str(shipment.dest_node_id))
            continue

        # Also check TLCs in the shipment against our contamination lookup
        shipment_tlcs = shipment.source_tlcs if shipment.source_tlcs else []
        is_contam, _ = check_shipment_contamination(shipment_tlcs)
        if is_contam:
            nodes_with_contamination.add(str(shipment.dest_node_id))

    for node_id, node in network.nodes.items():
        node_id_str = str(node_id)
        is_source = node_id_str in contaminated_node_ids
        received_contam = node_id_str in nodes_with_contamination

        ws_nodes.append([
            node_id_str,
            node.node_type.value,
            _get_node_name(network, node_id),
            node.location.city if node.location else "",
            node.location.state if node.location else "",
            "Yes" if is_source else "No",
            "Yes" if received_contam else "No",
        ])

    _style_header_row(ws_nodes, len(headers))
    _auto_column_width(ws_nodes)
    ws_nodes.freeze_panes = 'A2'

    # =========================================================================
    # Sheet 3: Traceability Records (CTEs/KDEs) - FSMA 204 Compliant
    # =========================================================================
    ws_trace = wb.create_sheet("Traceability_Records")

    headers = [
        "CTE Type",
        "Event ID",
        "Event Date",
        "TLC",
        "Product Category",
        "Quantity",
        "Unit of Measure",
        "Location ID",
        "Location Name",
        "Location Type",
        "IPS Location ID (Immediate Previous Source)",
        "IPS Location Name",
        "Source Reference (Farm Field/Area ID)",
        "Input TLCs (Deterministic)",
        "Input TLCs Probabilities (Probabilistic)",
        "ISR Location ID (Immediate Subsequent Recipient)",
        "ISR Location Name",
        "Contains Contaminated Product",
        "Contamination Probability",
    ]
    ws_trace.append(headers)

    # Track which TLCs were created at initial packing (at Packers)
    # These are the first "real" TLCs per FSMA 204
    initial_packing_events = {}  # tlc -> event data

    # First pass: identify INITIAL_PACKING events at Packers
    # In the simulation, farms create field-level IDs, packers receive and create TLCs
    for shipment in flow_sim.shipments:
        source_type = _get_node_type(network, shipment.source_node_id)
        dest_type = _get_node_type(network, shipment.dest_node_id)

        # Farm → Packer: No SHIP/RECEIVE CTE (no TLC exists yet)
        # But the Packer does INITIAL_PACKING when they receive from farm
        if source_type == "farm" and dest_type == "packer":
            for tlc in shipment.source_tlcs:
                if tlc not in initial_packing_events:
                    # Get farm field/area info from lot metadata
                    lot_meta = flow_sim.lot_metadata.get(tlc, {})
                    farm_field = lot_meta.get("growing_area", "")

                    # Look up actual contamination status from lot graph/seeder
                    is_contam, contam_prob = is_tlc_contaminated(tlc)

                    initial_packing_events[tlc] = {
                        "event_id": str(shipment.id),
                        "event_date": shipment.receive_date,
                        "tlc": tlc,
                        "packer_id": shipment.dest_node_id,
                        "farm_id": shipment.source_node_id,
                        "farm_field": farm_field,
                        "quantity": shipment.quantity.value / len(shipment.source_tlcs) if shipment.source_tlcs else shipment.quantity.value,
                        "unit": shipment.quantity.unit.value,
                        "product_category": shipment.product_category.value,
                        "contaminated": is_contam,
                        "contam_prob": contam_prob,
                    }

    # Write INITIAL_PACKING events (Packer creates first TLC)
    for tlc, event in initial_packing_events.items():
        ws_trace.append([
            "INITIAL_PACKING",
            event["event_id"],
            event["event_date"].isoformat() if event["event_date"] else "",
            event["tlc"],
            event["product_category"],
            event["quantity"],
            event["unit"],
            str(event["packer_id"]),
            _get_node_name(network, event["packer_id"]),
            "packer",
            str(event["farm_id"]),  # IPS is the farm
            _get_node_name(network, event["farm_id"]),
            event["farm_field"],  # Source reference (field/area ID from farm)
            "",  # No input TLCs - this is initial packing
            "",
            "",  # ISR not applicable
            "",
            "Yes" if event["contaminated"] else "No",
            event["contam_prob"],
        ])

    # Write SHIP and RECEIVE events (only from Packer onwards - when TLC exists)
    for shipment in flow_sim.shipments:
        source_type = _get_node_type(network, shipment.source_node_id)
        dest_type = _get_node_type(network, shipment.dest_node_id)

        # Skip Farm → Packer (no SHIP/RECEIVE CTE per FSMA 204)
        if source_type == "farm":
            continue

        # Check contamination status from TLCs in this shipment
        shipment_tlcs = shipment.source_tlcs if shipment.source_tlcs else []
        is_contam, contam_prob = check_shipment_contamination(shipment_tlcs)

        # SHIP event (from Packer, DC, etc.)
        ws_trace.append([
            "SHIP",
            str(shipment.id),
            shipment.ship_date.isoformat() if shipment.ship_date else "",
            ", ".join(shipment.source_tlcs) if shipment.source_tlcs else shipment.dest_tlc or "",
            shipment.product_category.value,
            shipment.quantity.value,
            shipment.quantity.unit.value,
            str(shipment.source_node_id),
            _get_node_name(network, shipment.source_node_id),
            source_type,
            "",  # IPS - not applicable for SHIP
            "",
            "",  # Source reference - not applicable for SHIP
            "",
            "",
            str(shipment.dest_node_id),  # ISR
            _get_node_name(network, shipment.dest_node_id),
            "Yes" if is_contam else "No",
            contam_prob,
        ])

        # RECEIVE event (at DC, Retailer, Deli)
        # Format probabilistic TLCs
        prob_tlcs_str = ""
        if shipment.tlc_probabilities:
            prob_tlcs_str = "; ".join([f"{tlc}: {prob:.3f}" for tlc, prob in shipment.tlc_probabilities.items()])

        ws_trace.append([
            "RECEIVE",
            str(shipment.id),
            shipment.receive_date.isoformat() if shipment.receive_date else "",
            shipment.dest_tlc or ", ".join(shipment.source_tlcs) if shipment.source_tlcs else "",
            shipment.product_category.value,
            shipment.quantity.value,
            shipment.quantity.unit.value,
            str(shipment.dest_node_id),
            _get_node_name(network, shipment.dest_node_id),
            dest_type,
            str(shipment.source_node_id),  # IPS
            _get_node_name(network, shipment.source_node_id),
            "",  # Source reference - not applicable for RECEIVE
            ", ".join(shipment.source_tlcs) if shipment.source_tlcs else "",  # Deterministic input TLCs
            prob_tlcs_str,  # Probabilistic input TLCs
            "",  # ISR - not applicable for RECEIVE
            "",
            "Yes" if is_contam else "No",
            contam_prob,
        ])

    # Write TRANSFORMATION events (at Delis/Processors that create new TLCs)
    if flow_sim.production_batches:
        for batch in flow_sim.production_batches:
            prob_inputs_str = ""
            if batch.input_tlc_probabilities:
                prob_inputs_str = "; ".join([f"{tlc}: {prob:.3f}" for tlc, prob in batch.input_tlc_probabilities.items()])

            # Check contamination of input TLCs
            input_tlcs = batch.input_tlcs if batch.input_tlcs else []
            is_contam, contam_prob = check_shipment_contamination(input_tlcs)

            ws_trace.append([
                "TRANSFORMATION",
                str(batch.id),
                batch.production_date.isoformat() if batch.production_date else "",
                batch.output_tlc,
                batch.output_product.value,
                batch.output_quantity.value,
                batch.output_quantity.unit.value,
                str(batch.processor_id),
                _get_node_name(network, batch.processor_id),
                _get_node_type(network, batch.processor_id),
                "",  # IPS - multiple inputs possible
                "",
                "",  # Source reference - not applicable
                ", ".join(batch.input_tlcs) if batch.input_tlcs else "",  # Deterministic input TLCs
                prob_inputs_str,  # Probabilistic input TLCs
                "",  # ISR - not applicable
                "",
                "Yes" if is_contam else "No",
                contam_prob,
            ])

    _style_header_row(ws_trace, len(headers))
    _auto_column_width(ws_trace)
    ws_trace.freeze_panes = 'A2'

    # =========================================================================
    # Sheet 4: Lot Codes (TLC Registry)
    # =========================================================================
    ws_lots = wb.create_sheet("Lot_Codes")

    headers = [
        "TLC",
        "Created At",
        "Created By Node ID",
        "Created By Node Name",
        "Created By Node Type",
        "Product Category",
        "Product Description",
        "Initial Quantity",
        "Unit of Measure",
        "Source TLCs (Deterministic)",
        "Source TLC Probabilities (Probabilistic)",
        "Is Contaminated",
        "Contamination Probability",
        "Contamination Source",
    ]
    ws_lots.append(headers)

    if lot_graph:
        for tlc, lot_record in lot_graph.lots.items():
            # Format probabilistic sources
            prob_sources_str = ""
            if lot_record.source_tlc_probabilities:
                prob_sources_str = "; ".join([
                    f"{src_tlc}: {prob:.3f}"
                    for src_tlc, prob in lot_record.source_tlc_probabilities.items()
                ])

            ws_lots.append([
                tlc,
                lot_record.created_at.isoformat(),
                str(lot_record.created_by_node_id),
                _get_node_name(network, lot_record.created_by_node_id),
                _get_node_type(network, lot_record.created_by_node_id),
                lot_record.product_category,
                lot_record.product_description,
                lot_record.initial_quantity_value,
                lot_record.initial_quantity_unit,
                ", ".join(lot_record.source_tlcs) if lot_record.source_tlcs else "(origin)",
                prob_sources_str if prob_sources_str else "(deterministic)",
                "Yes" if lot_record.is_contaminated else "No",
                lot_record.contamination_probability,
                lot_record.contamination_source or "",
            ])

    _style_header_row(ws_lots, len(headers))
    _auto_column_width(ws_lots)
    ws_lots.freeze_panes = 'A2'

    # =========================================================================
    # Sheet 5: Case Data
    # =========================================================================
    ws_cases = wb.create_sheet("Case_Data")

    headers = [
        "Case ID",
        "Consumer ID",
        "Case Status",
        "Pathogen",
        # Timeline
        "Exposure Date (Actual)",
        "Onset Date",
        "Report Date",
        # Exposure details
        "Exposure Location ID",
        "Exposure Location Name",
        "Exposure Product",
        # Interview data
        "Was Interviewed",
        "Interview Date",
        "Reported Exposure Location ID",
        "Reported Exposure Location Name",
        "Estimated Purchase Date (Patient Recall)",
        "Purchase Date Uncertainty Days",
        "Reported Location Matches Actual",
        # Clinical
        "Hospitalized",
        # Ground truth (for simulation validation)
        "Actual Contamination Source TLC",
    ]
    ws_cases.append(headers)

    for case in cases:
        reported_matches = ""
        if case.was_interviewed and case.reported_exposure_location_id:
            reported_matches = "Yes" if case.reported_exposure_location_id == case.exposure_location_id else "No"

        ws_cases.append([
            str(case.id),
            str(case.consumer_id),
            case.status.value,
            case.pathogen,
            case.exposure_date.isoformat() if case.exposure_date else "",
            case.onset_date.isoformat() if case.onset_date else "",
            case.report_date.isoformat() if case.report_date else "",
            str(case.exposure_location_id),
            case.exposure_location_name,
            case.exposure_product.value if case.exposure_product else "",
            "Yes" if case.was_interviewed else "No",
            case.interview_date.isoformat() if case.interview_date else "",
            str(case.reported_exposure_location_id) if case.reported_exposure_location_id else "",
            case.reported_exposure_location_name or "",
            case.estimated_purchase_date.isoformat() if case.estimated_purchase_date else "",
            case.purchase_date_uncertainty_days,
            reported_matches,
            "Yes" if case.hospitalized else "No",
            case.actual_contamination_source_tlc or "",
        ])

    _style_header_row(ws_cases, len(headers))
    _auto_column_width(ws_cases)
    ws_cases.freeze_panes = 'A2'

    # =========================================================================
    # Sheet 6: Contamination Events
    # =========================================================================
    ws_contam = wb.create_sheet("Contamination_Events")

    headers = [
        "Event ID",
        "Farm ID",
        "Farm Name",
        "Pathogen",
        "Start Date",
        "End Date",
        "Contamination Source",
        "Contamination Rate",
    ]
    ws_contam.append(headers)

    if seeder:
        for event in seeder.contamination_events:
            ws_contam.append([
                str(event.id),
                str(event.farm_id),
                event.farm_name,
                event.pathogen,
                event.start_date.isoformat() if event.start_date else "",
                event.end_date.isoformat() if event.end_date else "",
                event.contamination_source,
                event.contamination_rate,
            ])

    _style_header_row(ws_contam, len(headers))
    _auto_column_width(ws_contam)
    ws_contam.freeze_panes = 'A2'

    # =========================================================================
    # Sheet 7: Investigation Results
    # =========================================================================
    ws_invest = wb.create_sheet("Investigation_Results")

    # Get investigation results from the simulation result
    if run.result:
        det_scenario = run.result.get("scenarios", {}).get("deterministic", {})
        calc_scenario = run.result.get("scenarios", {}).get("calculated", {})
        metrics = run.result.get("metrics", {})

        ws_invest.append(["Investigation Results Comparison"])
        ws_invest.append([])
        ws_invest.append(["Metric", "Deterministic (Full Compliance)", "Probabilistic (Calculated)"])
        ws_invest.append(["Farms in Scope", det_scenario.get("farms_in_scope", ""), calc_scenario.get("farms_in_scope", "")])
        ws_invest.append(["TLCs in Scope", det_scenario.get("tlcs_in_scope", ""), calc_scenario.get("tlcs_in_scope", "")])
        ws_invest.append(["TLC Locations", det_scenario.get("tlcs_locations", ""), calc_scenario.get("tlcs_locations", "")])
        ws_invest.append(["Traceback Paths", det_scenario.get("traceback_paths", ""), calc_scenario.get("traceback_paths", "")])
        ws_invest.append(["Primary Suspect", det_scenario.get("primary_suspect", ""), calc_scenario.get("primary_suspect", "")])
        ws_invest.append(["Identification Outcome", det_scenario.get("identification_outcome", "").title(), calc_scenario.get("identification_outcome", "").title()])
        ws_invest.append(["Top Two Margin", det_scenario.get("top_two_margin", ""), calc_scenario.get("top_two_margin", "")])
        ws_invest.append(["Source Rank", det_scenario.get("source_rank", ""), calc_scenario.get("source_rank", "")])
        ws_invest.append([])
        ws_invest.append(["Expansion Metrics"])
        ws_invest.append(["Farm Scope Expansion", metrics.get("farm_scope_expansion", "")])
        ws_invest.append(["TLC Scope Expansion", metrics.get("tlc_scope_expansion", "")])
        ws_invest.append(["Path Expansion", metrics.get("path_expansion", "")])
        ws_invest.append([])
        ws_invest.append(["Actual Source Farm", det_scenario.get("actual_source", "")])

        # Farm probability distribution
        farm_probs = calc_scenario.get("farm_probabilities", {})
        if farm_probs:
            ws_invest.append([])
            ws_invest.append(["Farm Probability Distribution (Probabilistic Mode)"])
            ws_invest.append(["Farm Name", "Confidence Score"])
            for farm_name, prob in sorted(farm_probs.items(), key=lambda x: x[1], reverse=True):
                ws_invest.append([farm_name, prob])

    _auto_column_width(ws_invest)
    ws_invest.freeze_panes = 'A2'

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def get_export_filename(simulation_id: str) -> str:
    """Generate filename for export."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"simulation_export_{simulation_id[:8]}_{timestamp}.xlsx"
